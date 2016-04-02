#!/usr/bin/python
import sys, re
from openpyxl import load_workbook

pattern = re.compile("([0-9]+)+ +([A-Z]+)");
for i in range(1, len(sys.argv)):
    wb = load_workbook(sys.argv[i]);
    ws = wb.active;
    # print sys.argv[i]
    # print sys.argv[i], "has", len(ws.rows), "rows and", len(ws.columns), "columns."
    p = len(ws.columns);
    for j in range(2, len(ws.columns)+1):
        value = ws.cell(row=4, column=j).value;
        if value == None:
            p = j - 1;
            break;
        mo = pattern.match(value);
        quantity = int(mo.group(1));
        symbol = mo.group(2);

        print ("create table {}_SEK_Rates ("
               "day date primary key not null,"
               "rate double);").format(mo.group(2));
        print "insert into {}_SEK_Rates values ".format(symbol),
        for i in range(5, 1572):
            date = ws.cell(row=i, column=1).value;
            try:
                value = float(ws.cell(row=i, column=j).value)/quantity;
                print "(\"{}\", {}){}".format(date, value, "," if i < 1257 else ";");
            except:
                print "(\"{}\", {}){}".format(date, "NULL", "," if i < 1257 else ";");
